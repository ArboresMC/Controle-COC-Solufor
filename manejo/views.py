from datetime import date
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.cache import cache
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.views.generic import CreateView, ListView, TemplateView, UpdateView, View
from django.db import connection

import os
import tempfile
import uuid
import openpyxl

from .forms import EspecieForm, PropriedadeForm, InventarioEntradaForm, SaidaManejoForm
from .models import Especie, Propriedade, InventarioEntrada, SaidaManejo


class FMAccessMixin(LoginRequiredMixin):
    """Garante que só usuários com acesso a Manejo Florestal (FM) entrem nessas telas."""

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_manager or user.is_auditor:
            return super().dispatch(request, *args, **kwargs)
        if user.is_manejo_multi:
            return super().dispatch(request, *args, **kwargs)
        participant = getattr(user, 'participant', None)
        if not participant or not participant.ativo_fm:
            messages.error(request, 'Seu cadastro não tem acesso ao módulo de Manejo Florestal.')
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def is_multi_scope(self):
        """True quando o usuário pode operar mais de um participante FM
        (gestor/auditor veem o grupo inteiro; o 'mini-gestor' vê só os seus)."""
        user = self.request.user
        return user.is_manager or user.is_auditor or user.is_manejo_multi

    def get_allowed_participants(self):
        """Queryset de participantes FM que este usuário pode escolher/ver.
        Gestor/auditor: todo o grupo. Mini-gestor: só os seus. Participante
        comum: não usa este método (tem participante único e fixo)."""
        from participants.models import Participant
        user = self.request.user
        if user.is_manager or user.is_auditor:
            return Participant.objects.filter(ativo_fm=True, status='active').order_by('trade_name')
        if user.is_manejo_multi:
            return Participant.objects.filter(
                id__in=user.manejo_participant_ids(), status='active'
            ).order_by('trade_name')
        return Participant.objects.none()

    def get_participant(self):
        user = self.request.user
        if self.is_multi_scope():
            participant_id = self.request.GET.get('participant')
            if participant_id:
                allowed = self.get_allowed_participants()
                return allowed.filter(pk=participant_id).first()
            return None
        return getattr(user, 'participant', None)


class ManejoManagerRequiredMixin(LoginRequiredMixin):
    """Restringe acesso a telas administrativas do Manejo (gestão de dados,
    inativação de propriedades) apenas a gestores/auditores."""

    def dispatch(self, request, *args, **kwargs):
        if not (request.user.is_authenticated and (request.user.is_manager or request.user.is_auditor)):
            messages.error(request, 'Apenas o gestor pode acessar esta área.')
            return redirect('manejo_dashboard')
        return super().dispatch(request, *args, **kwargs)


def _build_participant_context(participant):
    """Monta o contexto de saldo/inventário de UM participante FM. Usado tanto
    pelo painel direto do participante quanto pelo drill-in do gestor."""
    entradas = InventarioEntrada.objects.filter(participant=participant).select_related('propriedade', 'especie').com_saldo()
    saldo_rows = []
    total_inicial = Decimal('0')
    total_saldo = Decimal('0')
    for entrada in entradas:
        saldo = entrada.saldo_disponivel_m3
        total_inicial += entrada.volume_m3
        total_saldo += saldo
        saldo_rows.append({
            'entrada': entrada,
            'propriedade': entrada.propriedade.nome,
            'especie': entrada.especie.nome,
            'volume_inicial': entrada.volume_m3,
            'volume_vendido': entrada.volume_vendido_m3,
            'saldo': saldo,
        })

    recent_saidas = SaidaManejo.objects.filter(participant=participant).exclude(
        documento='AJUSTE-HISTORICO'
    ).select_related('entrada__propriedade', 'entrada__especie').order_by('-data', '-id')[:10]

    return {
        'saldo_rows': saldo_rows,
        'total_inicial': total_inicial,
        'total_saldo': total_saldo,
        'total_propriedades': Propriedade.objects.filter(participant=participant, ativa=True).count(),
        'total_especies': Especie.objects.filter(participant=participant, ativo=True).count(),
        'recent_saidas': recent_saidas,
    }


def _build_chart_data(membros_qs, today):
    """Gráfico dos últimos 6 meses: entradas de inventário e saídas reais
    (exclui ajustes históricos) agregadas entre os membros informados.

    Otimizado para 2 queries agrupadas (em vez de 12 queries, uma por mês)
    e cacheado por 10 minutos — o mês corrente pode ficar levemente
    desatualizado nesse intervalo, mas evita recalcular o histórico
    completo a cada carregamento do painel."""
    from django.core.cache import cache
    from django.db.models import Count
    from django.db.models.functions import TruncMonth

    membros_ids = tuple(sorted(membros_qs.values_list('id', flat=True)))
    cache_key = f"fm_chart:{hash(membros_ids)}:{today.year}:{today.month}"
    cached = cache.get(cache_key)
    if cached:
        return cached

    meses_pt = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    months = []
    for i in range(5, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        months.append((y, m))

    inicio = date(months[0][0], months[0][1], 1)
    labels = [meses_pt[m - 1] for y, m in months]

    entradas_por_mes = {
        row['mes']: row['total']
        for row in InventarioEntrada.objects.filter(participant__in=membros_ids, data__gte=inicio)
        .annotate(mes=TruncMonth('data')).values('mes').annotate(total=Count('id')).values('mes', 'total')
    }
    saidas_por_mes = {
        row['mes']: row['total']
        for row in SaidaManejo.objects.filter(participant__in=membros_ids, data__gte=inicio)
        .exclude(documento='AJUSTE-HISTORICO')
        .annotate(mes=TruncMonth('data')).values('mes').annotate(total=Count('id')).values('mes', 'total')
    }

    entradas_data = [entradas_por_mes.get(date(y, m, 1), 0) for y, m in months]
    saidas_data = [saidas_por_mes.get(date(y, m, 1), 0) for y, m in months]

    result = {
        'fm_chart_labels': labels,
        'fm_chart_entradas': entradas_data,
        'fm_chart_saidas': saidas_data,
    }
    cache.set(cache_key, result, 600)  # 10 minutos
    return result


def _build_aggregate_context(membros_fm, request):
    """Monta o contexto agregado (totais + saldo por membro + gráfico) para
    um conjunto de participantes FM. Usado tanto pelo gestor (todos os
    membros do grupo) quanto pelo mini-gestor (só os participantes que ele
    opera)."""
    propriedades = Propriedade.objects.filter(participant__in=membros_fm, ativa=True)
    entradas = InventarioEntrada.objects.filter(participant__in=membros_fm).select_related(
        'propriedade', 'especie', 'participant'
    ).com_saldo()

    total_volume = Decimal('0')
    total_saldo = Decimal('0')
    por_membro = {}
    for entrada in entradas:
        saldo = entrada.saldo_disponivel_m3
        total_volume += entrada.volume_m3
        total_saldo += saldo
        pid = entrada.participant_id
        if pid not in por_membro:
            por_membro[pid] = {
                'nome': str(entrada.participant),
                'volume': Decimal('0'),
                'saldo': Decimal('0'),
                'propriedades': set(),
            }
        por_membro[pid]['volume'] += entrada.volume_m3
        por_membro[pid]['saldo'] += saldo
        por_membro[pid]['propriedades'].add(entrada.propriedade_id)

    membro_rows = sorted(
        [
            {
                'participant_id': pid,
                'nome': info['nome'],
                'propriedades_count': len(info['propriedades']),
                'volume': info['volume'],
                'saldo': info['saldo'],
            }
            for pid, info in por_membro.items()
        ],
        key=lambda r: r['nome']
    )

    from django.core.paginator import Paginator
    page = request.GET.get('page', 1)
    paginator = Paginator(membro_rows, 15)
    membro_rows_page = paginator.get_page(page)

    recent_saidas = SaidaManejo.objects.filter(participant__in=membros_fm).exclude(
        documento='AJUSTE-HISTORICO'
    ).select_related('entrada__propriedade', 'entrada__especie', 'participant').order_by('-data', '-id')[:10]

    ctx = {
        'is_aggregate': True,
        'total_membros': membros_fm.count() if hasattr(membros_fm, 'count') else len(membros_fm),
        'total_propriedades': propriedades.count(),
        'total_volume': total_volume,
        'total_saldo': total_saldo,
        'membro_rows': membro_rows_page,
        'recent_saidas': recent_saidas,
    }
    ctx.update(_build_chart_data(membros_fm, date.today()))
    return ctx


class ManejoDashboardView(FMAccessMixin, TemplateView):
    """Painel principal de Manejo Florestal.
    Gestor/auditor: visão agregada de TODOS os membros FM do grupo.
    Mini-gestor (multi-participante): visão agregada SÓ dos participantes que ele opera.
    Participante comum: visão direta dos seus próprios dados (sem agregação)."""
    template_name = 'manejo/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user

        if user.is_manager or user.is_auditor:
            from participants.models import Participant
            current_org = getattr(user, 'current_organization', None)
            membros_fm = Participant.objects.filter(ativo_fm=True, status='active')
            membros_fm = membros_fm.filter(organization=current_org) if current_org else Participant.objects.none()
            ctx.update(_build_aggregate_context(membros_fm, self.request))
            return ctx

        if user.is_manejo_multi:
            from participants.models import Participant
            membros_fm = Participant.objects.filter(id__in=user.manejo_participant_ids(), status='active')
            ctx['is_mini_gestor'] = True
            ctx.update(_build_aggregate_context(membros_fm, self.request))
            return ctx

        # Participante comum: mostra direto os próprios dados, sem agregação
        participant = getattr(user, 'participant', None)
        ctx['is_aggregate'] = False
        if not participant or not participant.ativo_fm:
            ctx['require_filter'] = True
            return ctx
        ctx.update(_build_participant_context(participant))
        ctx.update(_build_chart_data(
            type(participant).objects.filter(pk=participant.pk), date.today()
        ))
        return ctx


class ManejoParticipantDashboardView(FMAccessMixin, TemplateView):
    """Visão detalhada de UM participante FM específico — usada pelo gestor
    e pelo mini-gestor (multi-participante) para 'entrar' em um membro a
    partir do painel agregado."""
    template_name = 'manejo/participante_dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user

        if self.is_multi_scope():
            ctx['participants'] = self.get_allowed_participants()
            participant = self.get_participant()
            ctx['selected_participant'] = participant
            if not participant:
                ctx['require_filter'] = True
                return ctx
        else:
            participant = self.get_participant()

        ctx.update(_build_participant_context(participant))
        return ctx


# --- Espécies ---

class EspecieListView(FMAccessMixin, ListView):
    template_name = 'manejo/especie_list.html'
    context_object_name = 'especies'

    def get_queryset(self):
        participant = self.get_participant()
        if not participant:
            return Especie.objects.none()
        return Especie.objects.filter(participant=participant).order_by('nome')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.is_multi_scope():
            ctx['participants'] = self.get_allowed_participants()
        ctx['selected_participant'] = self.get_participant()
        return ctx


class EspecieCreateView(FMAccessMixin, CreateView):
    model = Especie
    form_class = EspecieForm
    template_name = 'manejo/especie_form.html'
    success_url = reverse_lazy('manejo_especie_list')

    def form_valid(self, form):
        form.instance.participant = self.get_participant()
        messages.success(self.request, 'Espécie cadastrada com sucesso.')
        return super().form_valid(form)


class EspecieUpdateView(FMAccessMixin, UpdateView):
    model = Especie
    form_class = EspecieForm
    template_name = 'manejo/especie_form.html'
    success_url = reverse_lazy('manejo_especie_list')

    def get_queryset(self):
        user = self.request.user
        if user.is_manager or user.is_auditor:
            from participants.models import Participant
            current_org = getattr(user, 'current_organization', None)
            membros_org = Participant.objects.filter(organization=current_org) if current_org else Participant.objects.none()
            return Especie.objects.filter(participant__in=membros_org)
        if user.is_manejo_multi:
            return Especie.objects.filter(participant_id__in=user.manejo_participant_ids())
        return Especie.objects.filter(participant=self.get_participant())

    def form_valid(self, form):
        messages.success(self.request, 'Espécie atualizada com sucesso.')
        return super().form_valid(form)


# --- Propriedades ---

class PropriedadeListView(FMAccessMixin, ListView):
    template_name = 'manejo/propriedade_list.html'
    context_object_name = 'propriedades'

    def get_queryset(self):
        participant = self.get_participant()
        if not participant:
            return Propriedade.objects.none()
        return Propriedade.objects.filter(participant=participant).order_by('nome')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.is_multi_scope():
            ctx['participants'] = self.get_allowed_participants()
        ctx['selected_participant'] = self.get_participant()
        return ctx


class PropriedadeCreateView(FMAccessMixin, CreateView):
    model = Propriedade
    form_class = PropriedadeForm
    template_name = 'manejo/propriedade_form.html'
    success_url = reverse_lazy('manejo_propriedade_list')

    def form_valid(self, form):
        form.instance.participant = self.get_participant()
        messages.success(self.request, 'Propriedade cadastrada com sucesso.')
        return super().form_valid(form)


class PropriedadeUpdateView(FMAccessMixin, UpdateView):
    model = Propriedade
    form_class = PropriedadeForm
    template_name = 'manejo/propriedade_form.html'
    success_url = reverse_lazy('manejo_propriedade_list')

    def get_queryset(self):
        user = self.request.user
        if user.is_manager or user.is_auditor:
            # Gestor/auditor pode editar qualquer propriedade do seu ambiente,
            # mesmo sem ?participant= na URL (ex: vindo direto da listagem).
            from participants.models import Participant
            current_org = getattr(user, 'current_organization', None)
            membros_org = Participant.objects.filter(organization=current_org) if current_org else Participant.objects.none()
            return Propriedade.objects.filter(participant__in=membros_org)
        if user.is_manejo_multi:
            return Propriedade.objects.filter(participant_id__in=user.manejo_participant_ids())
        return Propriedade.objects.filter(participant=self.get_participant())

    def form_valid(self, form):
        messages.success(self.request, 'Propriedade atualizada com sucesso.')
        return super().form_valid(form)


# --- Entradas de inventário ---

class InventarioEntradaListView(FMAccessMixin, ListView):
    template_name = 'manejo/entrada_list.html'
    context_object_name = 'entradas'
    paginate_by = 25

    def get_queryset(self):
        participant = self.get_participant()
        if not participant:
            return InventarioEntrada.objects.none()
        return InventarioEntrada.objects.filter(participant=participant).select_related('propriedade', 'especie').com_saldo()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.is_multi_scope():
            ctx['participants'] = self.get_allowed_participants()
        ctx['selected_participant'] = self.get_participant()
        return ctx


class InventarioEntradaCreateView(FMAccessMixin, CreateView):
    model = InventarioEntrada
    form_class = InventarioEntradaForm
    template_name = 'manejo/entrada_form.html'
    success_url = reverse_lazy('manejo_entrada_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['participant'] = self.get_participant()
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, 'Entrada de inventário registrada com sucesso.')
        return super().form_valid(form)


# --- Saídas ---

class SaidaManejoListView(FMAccessMixin, ListView):
    template_name = 'manejo/saida_list.html'
    context_object_name = 'saidas'
    paginate_by = 25

    def get_queryset(self):
        participant = self.get_participant()
        if not participant:
            return SaidaManejo.objects.none()
        return SaidaManejo.objects.filter(participant=participant).select_related(
            'entrada__propriedade', 'entrada__especie'
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.is_multi_scope():
            ctx['participants'] = self.get_allowed_participants()
        ctx['selected_participant'] = self.get_participant()
        return ctx


class SaidaManejoCreateView(FMAccessMixin, CreateView):
    model = SaidaManejo
    form_class = SaidaManejoForm
    template_name = 'manejo/saida_form.html'
    success_url = reverse_lazy('manejo_saida_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['participant'] = self.get_participant()
        return kwargs

    def form_valid(self, form):
        messages.success(self.request, 'Saída registrada com sucesso.')
        return super().form_valid(form)


# =============================================================================
# GESTÃO DE DADOS — Manejo Florestal (apenas gestor)
# =============================================================================

def _delete_entradas_sql(entrada_ids):
    """Exclui InventarioEntrada e, antes, as SaidaManejo vinculadas
    (entrada.PROTECT exige isso). Coleta os IDs dependentes antes de
    deletar, na mesma lógica já usada em transactions._delete_entries_sql."""
    if not entrada_ids:
        return 0, 0
    ids = list(entrada_ids)
    fmt = ','.join(['%s'] * len(ids))
    with connection.cursor() as c:
        c.execute(f"SELECT id FROM manejo_saidamanejo WHERE entrada_id IN ({fmt})", ids)
        saida_ids = [row[0] for row in c.fetchall()]

        if saida_ids:
            sfmt = ','.join(['%s'] * len(saida_ids))
            c.execute(f"DELETE FROM manejo_saidamanejo WHERE id IN ({sfmt})", saida_ids)

        c.execute(f"DELETE FROM manejo_inventarioentrada WHERE id IN ({fmt})", ids)

    return len(ids), len(saida_ids)


def _delete_saidas_sql(saida_ids):
    """Exclui SaidaManejo. Não há dependentes (nada referencia SaidaManejo via FK)."""
    if not saida_ids:
        return 0
    ids = list(saida_ids)
    fmt = ','.join(['%s'] * len(ids))
    with connection.cursor() as c:
        c.execute(f"DELETE FROM manejo_saidamanejo WHERE id IN ({fmt})", ids)
    return len(ids)


class ManejoDataManagementView(ManejoManagerRequiredMixin, TemplateView):
    template_name = 'manejo/data_management.html'

    def get_context_data(self, **kwargs):
        from django.core.paginator import Paginator
        from participants.models import Participant
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        current_org = getattr(user, 'current_organization', None)

        participant_id = self.request.GET.get('participant')
        record_type = self.request.GET.get('type', 'propriedades')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        page = self.request.GET.get('page', 1)

        participants = Participant.objects.filter(
            ativo_fm=True, status='active', organization=current_org
        ).order_by('trade_name', 'legal_name') if current_org else Participant.objects.none()

        selected_participant = participants.filter(pk=participant_id).first() if participant_id else None

        records_page = None
        total_count = 0
        if selected_participant:
            records_qs = None
            if record_type == 'propriedades':
                records_qs = Propriedade.objects.filter(
                    participant=selected_participant, ativa=True
                ).order_by('nome')
            elif record_type == 'entradas':
                records_qs = InventarioEntrada.objects.filter(
                    participant=selected_participant
                ).select_related('propriedade', 'especie').com_saldo().order_by('-data', '-id')
            elif record_type == 'saidas':
                records_qs = SaidaManejo.objects.filter(
                    participant=selected_participant
                ).select_related('entrada__propriedade', 'entrada__especie').order_by('-data', '-id')

            if records_qs is not None:
                if record_type in ('entradas', 'saidas'):
                    if date_from:
                        try:
                            from datetime import datetime
                            records_qs = records_qs.filter(data__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
                        except ValueError:
                            pass
                    if date_to:
                        try:
                            from datetime import datetime
                            records_qs = records_qs.filter(data__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
                        except ValueError:
                            pass

                total_count = records_qs.count()
                paginator = Paginator(records_qs, 50)
                records_page = paginator.get_page(page)

        ctx.update({
            'participants': participants,
            'selected_participant': selected_participant,
            'record_type': record_type,
            'records': records_page,
            'total_count': total_count,
            'date_from': date_from,
            'date_to': date_to,
        })
        return ctx


class ManejoDataDeactivateView(ManejoManagerRequiredMixin, View):
    """Inativa (não exclui) Propriedade(s) — individual ou em lote.
    Propriedade tem on_delete=PROTECT a partir de InventarioEntrada,
    então exclusão de verdade quebraria o histórico FSC; inativar é o
    caminho correto para "encerrar" uma área sem perder rastreabilidade."""

    def post(self, request, *args, **kwargs):
        participant_id = request.POST.get('participant_id')
        action = request.POST.get('action')

        if action == 'deactivate_selected':
            selected_ids = request.POST.getlist('selected_ids')
            ids = [int(i) for i in selected_ids if i.isdigit()]
            if not ids:
                messages.warning(request, 'Nenhuma propriedade selecionada.')
            else:
                count = Propriedade.objects.filter(pk__in=ids).update(ativa=False)
                messages.success(request, f'{count} propriedade(s) inativada(s) com sucesso.')

        elif action == 'deactivate_single':
            record_id = request.POST.get('record_id')
            if record_id and record_id.isdigit():
                updated = Propriedade.objects.filter(pk=int(record_id)).update(ativa=False)
                if updated:
                    messages.success(request, 'Propriedade inativada com sucesso.')
                else:
                    messages.error(request, 'Propriedade não encontrada.')

        return redirect(f'/manejo/gestor/dados/?participant={participant_id}&type=propriedades')


class ManejoDataDeleteView(ManejoManagerRequiredMixin, View):
    """Exclusão (em lote ou total) de Entradas de Inventário ou Saídas de Manejo."""

    def post(self, request, *args, **kwargs):
        record_type = request.POST.get('record_type')
        participant_id = request.POST.get('participant_id')
        action = request.POST.get('action')

        from participants.models import Participant
        participant = Participant.objects.filter(pk=participant_id, ativo_fm=True).first() if participant_id else None
        if not participant:
            messages.error(request, 'Participante não encontrado.')
            return redirect('manejo_data_management')

        if action == 'delete_selected':
            selected_ids = request.POST.getlist('selected_ids')
            ids = [int(i) for i in selected_ids if i.isdigit()]
            if not ids:
                messages.warning(request, 'Nenhum registro selecionado.')
            elif record_type == 'entradas':
                count, saidas_count = _delete_entradas_sql(ids)
                extra = f' (e {saidas_count} saída(s) vinculada(s))' if saidas_count else ''
                messages.success(request, f'{count} entrada(s) excluída(s) com sucesso{extra}.')
            elif record_type == 'saidas':
                count = _delete_saidas_sql(ids)
                messages.success(request, f'{count} saída(s) excluída(s) com sucesso.')

        elif action == 'delete_all':
            if record_type == 'entradas':
                ids = list(InventarioEntrada.objects.filter(participant=participant).values_list('id', flat=True))
                count, saidas_count = _delete_entradas_sql(ids)
                extra = f' (e {saidas_count} saída(s) vinculada(s))' if saidas_count else ''
                messages.success(request, f'Todas as {count} entradas de {participant} foram excluídas{extra}.')
            elif record_type == 'saidas':
                ids = list(SaidaManejo.objects.filter(participant=participant).values_list('id', flat=True))
                count = _delete_saidas_sql(ids)
                messages.success(request, f'Todas as {count} saídas de {participant} foram excluídas.')

        return redirect(f'/manejo/gestor/dados/?participant={participant_id}&type={record_type}')


class ManejoDataDeleteSingleView(ManejoManagerRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        record_type = request.POST.get('record_type')
        record_id = request.POST.get('record_id')
        participant_id = request.POST.get('participant_id')

        if not record_id or not record_id.isdigit():
            messages.error(request, 'Registro inválido.')
            return redirect('manejo_data_management')

        rid = int(record_id)
        if record_type == 'entradas':
            count, saidas_count = _delete_entradas_sql([rid])
            extra = f' (e {saidas_count} saída(s) vinculada(s))' if saidas_count else ''
            messages.success(request, f'Entrada excluída com sucesso{extra}.')
        elif record_type == 'saidas':
            _delete_saidas_sql([rid])
            messages.success(request, 'Saída excluída com sucesso.')

        return redirect(f'/manejo/gestor/dados/?participant={participant_id}&type={record_type}')


# =============================================================================
# IMPORTAÇÃO EM LOTE — Manejo Florestal (Entradas e Saídas via planilha)
# =============================================================================

class ManejoImportTemplateDownloadView(LoginRequiredMixin, View):
    """Gera (com cache) o modelo de planilha Excel para importação em lote
    de Entradas (inventário) e Saídas de Manejo Florestal."""
    CACHE_KEY = 'import_template_manejo_xlsx_v1'
    CACHE_TTL = 60 * 60 * 24  # 24 horas

    def get(self, request, *args, **kwargs):
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        cached = cache.get(self.CACHE_KEY)
        if cached:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=modelo_importacao_manejo.xlsx'
            response.write(cached)
            return response

        COR_HEADER  = "1A6B3C"
        COR_OBRIG   = "E8F5E9"
        COR_INFO    = "FFF8E1"
        COR_TITULO  = "0F4A28"
        borda = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        def hdr(ws, row, col, text, width=18):
            c = ws.cell(row=row, column=col, value=text)
            c.font = Font(bold=True, color="FFFFFF", name='Arial', size=10)
            c.fill = PatternFill("solid", fgColor=COR_HEADER)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = borda
            ws.column_dimensions[get_column_letter(col)].width = width

        def cell(ws, row, col, value=None, bg=COR_OBRIG, italic=False):
            c = ws.cell(row=row, column=col, value=value)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(italic=italic, name='Arial', size=10)
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = borda

        def title(ws, row, text, ncols, bg=COR_TITULO):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            c = ws.cell(row=row, column=1, value=text)
            c.font = Font(bold=True, color="FFFFFF", name='Arial', size=11)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 28

        def info(ws, row, col1, col2, t, d, bg="FFFFFF"):
            ws.row_dimensions[row].height = 38
            for col, val in ((col1, t), (col2, d)):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(bold=(col == col1), name='Arial', size=10)
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(wrap_text=True, vertical='center')
                c.border = borda

        wb = openpyxl.Workbook()

        # ── Leia Antes ──────────────────────────────────────────
        wi = wb.active
        wi.title = "Leia Antes"
        wi.sheet_view.showGridLines = False
        wi.column_dimensions['A'].width = 34
        wi.column_dimensions['B'].width = 60

        title(wi, 1, "MODELO DE IMPORTAÇÃO — MANEJO FLORESTAL", 2)
        title(wi, 2, "FLUXO DE PREENCHIMENTO", 2, bg="1F7A4D")
        info(wi, 3, 1, 2, "1. Cadastre Propriedades e Espécies antes",
             "A planilha NÃO cria propriedades nem espécies novas. Cadastre-as no sistema antes de importar (evita duplicidade e erros de digitação).", COR_INFO)
        info(wi, 4, 1, 2, "2. Preencha Entradas",
             "Um inventário por combinação Propriedade + Espécie. Se já existir uma entrada para essa combinação, a importação será bloqueada — edite a entrada existente em vez de duplicar.", COR_OBRIG)
        info(wi, 5, 1, 2, "3. Preencha Saidas",
             "Informe Propriedade + Espécie exatamente como cadastradas. O sistema localiza a entrada de inventário correspondente automaticamente e debita o saldo.", COR_OBRIG)
        info(wi, 6, 1, 2, "4. Saldo insuficiente",
             "Se o volume da saída for maior que o saldo disponível da propriedade+espécie, a linha é rejeitada com erro — corrija o volume ou a entrada antes de tentar novamente.", COR_INFO)
        title(wi, 7, "LEGENDA DE CORES", 2, bg="1F7A4D")
        info(wi, 8, 1, 2, "Verde claro → Obrigatório", "Preencha antes de importar.", COR_OBRIG)
        info(wi, 9, 1, 2, "Amarelo     → Instrução",   "Leia com atenção antes de preencher.", COR_INFO)
        title(wi, 10, "UNIDADES ACEITAS", 2, bg="1F7A4D")
        info(wi, 11, 1, 2, "m3",  "Metro cúbico.")
        info(wi, 12, 1, 2, "ton", "Tonelada.")
        info(wi, 13, 1, 2, "st",  "Estéreo (st).")

        # ── Entradas ─────────────────────────────────────────────
        we = wb.create_sheet("Entradas")
        we.sheet_view.showGridLines = False
        we.freeze_panes = "A2"
        we.row_dimensions[1].height = 30
        cols_e = [("propriedade", 24), ("especie", 22), ("data", 16), ("documento", 22),
                  ("volume", 14), ("unidade", 12), ("observacoes", 30)]
        for i, (n, w) in enumerate(cols_e, 1):
            hdr(we, 1, i, n, w)
        ex_e = ["Xadrez", "Pinus taeda", "2026-03-18", "Laudo-001", 282944.77, "m3", "Inventário inicial"]
        for i, v in enumerate(ex_e, 1):
            cell(we, 2, i, v, COR_OBRIG)
        for row in range(3, 52):
            we.row_dimensions[row].height = 18
            for col in range(1, 8):
                cell(we, row, col, None, COR_OBRIG)

        # ── Saidas ───────────────────────────────────────────────
        wsai = wb.create_sheet("Saidas")
        wsai.sheet_view.showGridLines = False
        wsai.freeze_panes = "A3"
        wsai.row_dimensions[1].height = 34
        wsai.merge_cells('A1:H1')
        c = wsai.cell(row=1, column=1,
                      value='⚠ propriedade + especie devem ser EXATAMENTE iguais às cadastradas. O sistema localiza a entrada de inventário e debita o saldo automaticamente.')
        c.font = Font(bold=True, name='Arial', size=10, color="7B3F00")
        c.fill = PatternFill("solid", fgColor=COR_INFO)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wsai.row_dimensions[2].height = 30
        cols_s = [("propriedade", 24), ("especie", 22), ("data", 16), ("documento", 22),
                  ("cliente", 26), ("declaracao_fsc", 16), ("volume", 14), ("unidade", 12), ("observacoes", 30)]
        for i, (n, w) in enumerate(cols_s, 1):
            hdr(wsai, 2, i, n, w)
        ex_s = ["Xadrez", "Pinus taeda", "2026-03-20", "NF-0010", "Tramontina", "Sim", 40, "m3", "Baixa parcial"]
        for i, v in enumerate(ex_s, 1):
            cell(wsai, 3, i, v, COR_OBRIG)
        for row in range(4, 52):
            wsai.row_dimensions[row].height = 18
            for col in range(1, 10):
                cell(wsai, row, col, None, COR_OBRIG)

        wb.active = wi
        buffer = BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()
        cache.set(self.CACHE_KEY, xlsx_bytes, self.CACHE_TTL)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=modelo_importacao_manejo.xlsx'
        response.write(xlsx_bytes)
        return response


class ManejoImportWorkbookView(FMAccessMixin, View):
    """Upload, validação e confirmação de importação em lote de Entradas e
    Saídas de Manejo Florestal. Segue o mesmo fluxo de duas etapas do
    importador de Cadeia de Custódia: valida sem gravar, e só persiste após
    confirmação explícita do usuário (sem precisar reenviar o arquivo)."""
    template_name = 'manejo/import_workbook.html'

    def _render(self, request, preview=None, errors=None, summary=None,
                validated_token=None, validated_filename=None, selected_participant=None):
        from reports.services import humanize_import_errors
        ctx = {
            'preview': preview or {},
            'preview_errors': humanize_import_errors(errors or []),
            'summary': summary or {},
            'validated_token': validated_token,
            'validated_filename': validated_filename,
        }
        if self.is_multi_scope():
            ctx['participants'] = self.get_allowed_participants()
        ctx['selected_participant'] = selected_participant or self.get_participant()
        return render(request, self.template_name, ctx)

    def get(self, request, *args, **kwargs):
        return self._render(request)

    def _save_temp_file(self, request, uploaded_file):
        token = str(uuid.uuid4())
        uploaded_file.seek(0)
        content = uploaded_file.read()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', prefix=f'import_manejo_{token}_')
        tmp.write(content)
        tmp.close()
        session_key = f'import_manejo_tmp_{token}'
        request.session[session_key] = {
            'path': tmp.name,
            'filename': getattr(uploaded_file, 'name', 'planilha.xlsx'),
        }
        return token

    def _load_temp_file(self, request, token):
        session_key = f'import_manejo_tmp_{token}'
        data = request.session.get(session_key)
        if not data:
            return None, None
        path = data.get('path')
        filename = data.get('filename', 'planilha.xlsx')
        if not path or not os.path.exists(path):
            return None, None
        with open(path, 'rb') as f:
            content = f.read()
        return content, filename

    def _cleanup_temp_file(self, request, token):
        session_key = f'import_manejo_tmp_{token}'
        data = request.session.pop(session_key, None)
        if data and data.get('path'):
            try:
                os.unlink(data['path'])
            except OSError:
                pass

    def post(self, request, *args, **kwargs):
        from io import BytesIO
        from .import_services import build_manejo_import_preview
        action = request.POST.get('action') or 'validate'

        participant_id = request.POST.get('participant_id') or request.GET.get('participant')
        if self.is_multi_scope():
            allowed = self.get_allowed_participants()
            participant = allowed.filter(pk=participant_id).first() if participant_id else None
        else:
            participant = getattr(request.user, 'participant', None)

        if not participant:
            messages.error(request, 'Selecione um participante de Manejo válido.')
            return self._render(request)

        # ── Confirmação sem re-upload ─────────────────────────────
        if action == 'confirm':
            token = request.POST.get('validated_token', '')
            file_bytes, filename = self._load_temp_file(request, token)
            if not file_bytes:
                messages.error(request, 'Sessão expirada. Por favor, valide a planilha novamente.')
                return self._render(request, selected_participant=participant)

            workbook = openpyxl.load_workbook(BytesIO(file_bytes))
            self._cleanup_temp_file(request, token)
            summary, errors, preview = build_manejo_import_preview(workbook, participant, request.user, persist=True)
            if errors:
                messages.error(request, 'A importação encontrou inconsistências e não foi concluída. Revise a planilha.')
                return self._render(request, preview=preview, errors=errors, summary=summary, selected_participant=participant)
            messages.success(
                request,
                f"Importação concluída. Entradas: {summary.get('entradas', 0)}, saídas: {summary.get('saidas', 0)}."
            )
            redirect_url = reverse('manejo_dashboard')
            if self.is_multi_scope():
                redirect_url += f'?participant={participant.id}'
            return redirect(redirect_url)

        # ── Validação ──────────────────────────────────────────────
        uploaded_file = request.FILES.get('workbook')
        if not uploaded_file:
            messages.error(request, 'Selecione um arquivo de planilha (.xlsx).')
            return self._render(request, selected_participant=participant)

        workbook = openpyxl.load_workbook(uploaded_file)
        summary, errors, preview = build_manejo_import_preview(workbook, participant, request.user, persist=False)

        validated_token = None
        validated_filename = None
        if not errors:
            validated_token = self._save_temp_file(request, uploaded_file)
            validated_filename = getattr(uploaded_file, 'name', 'planilha.xlsx')

        if errors:
            messages.warning(request, f'Validação concluída com {len(errors)} inconsistência(s). Corrija a planilha antes de importar.')
        else:
            messages.success(request, 'Validação concluída sem inconsistências. Confirme para importar.')

        return self._render(
            request, preview=preview, errors=errors, summary=summary,
            validated_token=validated_token, validated_filename=validated_filename,
            selected_participant=participant,
        )
