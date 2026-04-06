from decimal import Decimal
from unicodedata import normalize

from django.conf import settings
from django.core.cache import cache
from django.core.paginator import Paginator

import os
import tempfile
import uuid
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.urls import reverse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views import View

from catalog.models import Counterparty, Product
from participants.models import Participant
from transactions.models import EntryRecord, SaleRecord, TransformationRecord
from transactions.performance import build_cache_key
from transactions.services import (
    build_traceability_rows,
    calculate_target_from_source,
    convert_to_base,
    get_entry_balance_rows,
    get_transformation_rule,
    reallocate_sale,
    reallocate_transformation_sources,
    sync_entry_lot,
    sync_transformation_target_lot,
)
from .forms import ImportWorkbookForm
from .models import ImportJob
from .services import (
build_import_error_workbook,
    build_import_preview,
    count_workbook_rows,
    humanize_import_errors,

)


def _serialize_json_safe(value):
    if isinstance(value, dict):
        return {key: _serialize_json_safe(val) for key, val in value.items()}
    if isinstance(value, (list, tuple)):
        return [_serialize_json_safe(item) for item in value]
    if hasattr(value, 'isoformat'):
        try:
            return value.isoformat()
        except Exception:
            pass
    if isinstance(value, Decimal):
        return str(value)
    return value



class ManagerRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_manager



class ConsolidatedExcelReportView(ManagerRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Entradas'
        ws1.append(['Data', 'Participante', 'Documento', 'Fornecedor', 'Produto', 'Qtd informada', 'Unidade informada', 'Qtd base', 'Unidade base', 'Declaração FSC', 'Status'])
        for obj in EntryRecord.objects.select_related('participant', 'supplier', 'product'):
            ws1.append([obj.movement_date.strftime('%d/%m/%Y'), str(obj.participant), obj.document_number, str(obj.supplier), str(obj.product), float(obj.quantity), obj.movement_unit, float(obj.quantity_base), obj.unit_snapshot, obj.fsc_claim, obj.get_status_display()])

        ws2 = wb.create_sheet('Saídas')
        ws2.append(['Data', 'Participante', 'Documento', 'Cliente', 'Produto', 'Qtd informada', 'Unidade informada', 'Qtd base', 'Unidade base', 'Declaração FSC', 'Status'])
        for obj in SaleRecord.objects.select_related('participant', 'customer', 'product'):
            ws2.append([obj.movement_date.strftime('%d/%m/%Y'), str(obj.participant), obj.document_number, str(obj.customer), str(obj.product), float(obj.quantity), obj.movement_unit, float(obj.quantity_base), obj.unit_snapshot, obj.fsc_claim, obj.get_status_display()])

        ws3 = wb.create_sheet('Transformações')
        ws3.append(['Data', 'Participante', 'Documento', 'Cliente final', 'Fornecedor origem', 'Declaração FSC', 'Produto origem', 'Qtd origem', 'Unidade origem', 'Qtd origem base', 'Produto destino', 'Qtd destino base', 'Unidade base destino', 'Fator'])
        for obj in TransformationRecord.objects.select_related('participant', 'source_product', 'target_product'):
            ws3.append([obj.movement_date.strftime('%d/%m/%Y'), str(obj.participant), obj.document_number, str(obj.customer) if obj.customer else '', str(obj.supplier) if obj.supplier else '', obj.fsc_claim, str(obj.source_product), float(obj.source_quantity), obj.source_unit, float(obj.source_quantity_base), str(obj.target_product), float(obj.target_quantity_base), obj.target_unit_snapshot, float(obj.yield_factor_snapshot)])

        ws4 = wb.create_sheet('Rastreabilidade')
        ws4.append(['Participante', 'Produto', 'Tipo de uso', 'Uso', 'Data uso', 'Fornecedor origem', 'Cliente / destino', 'Origem consumida', 'Data origem', 'Quantidade', 'Unidade'])
        for row in build_traceability_rows():
            ws4.append([str(row['participant']), str(row['product']), row['use_type'], row['use_label'], row['use_date'].strftime('%d/%m/%Y'), row['supplier'], row['counterparty'], row['source_label'], row['source_date'].strftime('%d/%m/%Y'), float(row['quantity']), row['unit']])

        ws5 = wb.create_sheet('Saldo por entrada')
        ws5.append(['Participante', 'Data entrada', 'Documento entrada', 'Fornecedor', 'Produto', 'Qtd entrada', 'Qtd vendida', 'Qtd transformada', 'Saldo remanescente', 'Unidade', 'Clientes atendidos'])
        for row in get_entry_balance_rows():
            ws5.append([str(row['participant']), row['movement_date'].strftime('%d/%m/%Y'), row['entry'].document_number, str(row['supplier']) if row['supplier'] else '', str(row['product']), float(row['quantity_total']), float(row['quantity_sold']), float(row['quantity_transformed']), float(row['quantity_remaining']), row['unit'], row['customers']])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f'relatorio_fsc_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


class AuditPdfReportView(ManagerRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        participant_id = request.GET.get('participant')
        participant = Participant.objects.filter(pk=participant_id).first() if participant_id else None
        rows = build_traceability_rows(participant=participant)[:80]
        entry_balances = get_entry_balance_rows(participant=participant)[:80]

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=relatorio_auditoria_fsc_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()
        story = [
            Paragraph('Solufor Soluções Florestais', styles['Title']),
            Paragraph('Relatório de auditoria FSC — rastreabilidade e saldo por entrada', styles['Heading2']),
            Spacer(1, 12),
        ]
        if participant:
            story.append(Paragraph(f'Participante filtrado: {participant}', styles['Normal']))
            story.append(Spacer(1, 10))

        story.append(Paragraph('Rastreabilidade fornecedor → cliente', styles['Heading3']))
        trace_table = [['Participante', 'Produto', 'Uso', 'Fornecedor origem', 'Cliente / destino', 'Qtd']]
        for row in rows:
            trace_table.append([
                str(row['participant']), str(row['product']), row['use_label'], row['supplier'], row['counterparty'],
                f"{row['quantity']} {row['unit']}"
            ])
        table = Table(trace_table, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#14532d')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        story.append(table)
        story.append(Spacer(1, 14))

        story.append(Paragraph('Saldo por entrada / lote', styles['Heading3']))
        bal_table = [['Participante', 'Entrada', 'Fornecedor', 'Produto', 'Qtd entrada', 'Saldo remanescente']]
        for row in entry_balances:
            bal_table.append([
                str(row['participant']), row['entry'].document_number, str(row['supplier'] or ''), str(row['product']),
                f"{row['quantity_total']} {row['unit']}", f"{row['quantity_remaining']} {row['unit']}"
            ])
        table2 = Table(bal_table, repeatRows=1)
        table2.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        story.append(table2)
        doc.build(story)
        return response


class TraceabilityReportView(ManagerRequiredMixin, View):
    template_name = 'reports/traceability_report.html'

    def get(self, request, *args, **kwargs):
        participant_id = request.GET.get('participant')
        participant = Participant.objects.filter(pk=participant_id).first() if participant_id else None
        current_org = getattr(request.user, 'current_organization', None)
        scope_type = 'organization' if (request.user.is_manager or request.user.is_auditor) else 'participant'
        scope_id = getattr(current_org, 'id', None) if scope_type == 'organization' else getattr(getattr(request.user, 'participant', None), 'id', None)
        cache_extra = f"traceability:{participant_id or 'all'}"
        cache_key = build_cache_key('report', scope_type, scope_id, extra=cache_extra)
        cached = cache.get(cache_key)
        if cached is None:
            rows = build_traceability_rows(participant=participant)
            entry_balances = get_entry_balance_rows(participant=participant)
            participants = list(Participant.objects.filter(status='active'))
            cached = {'rows': rows, 'entry_balances': entry_balances, 'participants': participants}
            cache.set(cache_key, cached, settings.CACHE_TTL_TRACEABILITY)

        rows_paginator = Paginator(cached['rows'], 80)
        balances_paginator = Paginator(cached['entry_balances'], 80)
        rows_page = rows_paginator.get_page(request.GET.get('page'))
        balances_page = balances_paginator.get_page(request.GET.get('balance_page'))
        query_for_rows = request.GET.copy()
        query_for_rows.pop('page', None)
        query_for_balances = request.GET.copy()
        query_for_balances.pop('balance_page', None)

        return render(request, self.template_name, {
            'rows': rows_page,
            'entry_balances': balances_page,
            'participants': cached['participants'],
            'selected_participant': participant,
            'rows_page_obj': rows_page,
            'balances_page_obj': balances_page,
            'rows_query_string': query_for_rows.urlencode(),
            'balances_query_string': query_for_balances.urlencode(),
        })


class ImportTemplateDownloadView(LoginRequiredMixin, View):
    CACHE_KEY = 'import_template_xlsx_v2'
    CACHE_TTL = 60 * 60 * 24  # 24 horas

    def get(self, request, *args, **kwargs):
        from io import BytesIO
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Serve do cache se disponível — evita regenerar a cada clique
        cached = cache.get(self.CACHE_KEY)
        if cached:
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=modelo_importacao_fsc.xlsx'
            response.write(cached)
            return response

        COR_HEADER  = "005B78"
        COR_OBRIG   = "E6F4F8"
        COR_AUTO    = "E8F5E9"
        COR_INFO    = "FFF8E1"
        COR_TITULO  = "003D52"
        borda = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        def hdr(ws, row, col, text, width=18, auto=False):
            c = ws.cell(row=row, column=col, value=text)
            c.font = Font(bold=True, color="FFFFFF", name='Arial', size=10)
            c.fill = PatternFill("solid", fgColor="1A7A5E" if auto else COR_HEADER)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = borda
            ws.column_dimensions[get_column_letter(col)].width = width

        def cell(ws, row, col, value=None, bg=COR_OBRIG, italic=False):
            c = ws.cell(row=row, column=col, value=value)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(italic=italic, name='Arial', size=10)
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.border = borda

        def title(ws, row, text, ncols, bg=COR_TITULO):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            c = ws.cell(row=row, column=1, value=text)
            c.font = Font(bold=True, color="FFFFFF", name='Arial', size=11)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 28

        def info(ws, row, col1, col2, t, d, bg="FFFFFF"):
            ws.row_dimensions[row].height = 38
            for col, val in ((col1, t), (col2, d)):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(bold=(col == col1), name='Arial', size=10)
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(wrap_text=True, vertical='center')
                c.border = borda

        wb = openpyxl.Workbook()

        # ── Leia Antes ──────────────────────────────────────────
        wi = wb.active
        wi.title = "Leia Antes"
        wi.sheet_view.showGridLines = False
        wi.column_dimensions['A'].width = 34
        wi.column_dimensions['B'].width = 60

        title(wi, 1, "MODELO DE IMPORTAÇÃO FSC – SOLUFOR", 2)
        title(wi, 2, "FLUXO DE PREENCHIMENTO", 2, bg="006E91")
        info(wi, 3, 1, 2, "1. Preencha Entradas",
             "Informe: data, número do documento (ex: NF-0001), fornecedor, produto, quantidade, unidade e declaração FSC.", COR_OBRIG)
        info(wi, 4, 1, 2, "2. Preencha Saidas",
             "No campo documento_origem informe o número exato do documento da entrada que gerou este lote (ex: NF-0001). O sistema localiza o lote correto automaticamente.", COR_AUTO)
        info(wi, 5, 1, 2, "3. Preencha Transformacoes",
             "Igual à saída: informe documento_origem com o número do documento da entrada. O sistema identifica o produto de origem e calcula o consumo pela regra de rendimento.", COR_AUTO)
        info(wi, 6, 1, 2, "4. documento_origem em branco",
             "Se deixar em branco, o sistema aloca por FIFO (ordem de chegada). Sempre prefira preencher para garantir rastreabilidade FSC completa.", COR_INFO)
        title(wi, 7, "LEGENDA DE CORES", 2, bg="006E91")
        info(wi, 8,  1, 2, "Azul claro  → Obrigatório",    "Preencha antes de importar.", COR_OBRIG)
        info(wi, 9,  1, 2, "Verde claro → Informativo",    "Preenchido automaticamente pelo sistema. NÃO precisa preencher.", COR_AUTO)
        info(wi, 10, 1, 2, "Amarelo     → Instrução",      "Leia com atenção antes de preencher.", COR_INFO)
        title(wi, 11, "UNIDADES ACEITAS", 2, bg="006E91")
        info(wi, 12, 1, 2, "m3", "Metro cúbico — volumes de toras e madeira serrada.")
        info(wi, 13, 1, 2, "kg", "Quilograma.")
        info(wi, 14, 1, 2, "t",  "Tonelada métrica.")
        info(wi, 15, 1, 2, "un", "Unidade.")
        title(wi, 16, "DECLARAÇÕES FSC COMUNS", 2, bg="006E91")
        info(wi, 17, 1, 2, "FSC 100%",           "Toda a madeira é de origem FSC certificada.")
        info(wi, 18, 1, 2, "FSC Mix",             "Mistura de madeira FSC e controlada.")
        info(wi, 19, 1, 2, "FSC Controlled Wood", "Madeira de origem controlada FSC.")

        # ── Entradas ─────────────────────────────────────────────
        we = wb.create_sheet("Entradas")
        we.sheet_view.showGridLines = False
        we.freeze_panes = "A2"
        we.row_dimensions[1].height = 30
        cols_e = [("data",16),("documento",22),("fornecedor",28),("produto",24),
                  ("quantidade",14),("unidade",12),("declaracao_fsc",20),("lote",16),("observacoes",30)]
        for i, (n, w) in enumerate(cols_e, 1):
            hdr(we, 1, i, n, w)
        ex_e = ["2026-03-18", "NF-0001", "Berneck S/A", "Toras e Toretes", 45, "t", "FSC 100%", "L001", "Compra inicial de toras"]
        for i, v in enumerate(ex_e, 1):
            cell(we, 2, i, v, COR_OBRIG)
        for row in range(3, 52):
            we.row_dimensions[row].height = 18
            for col in range(1, 10):
                cell(we, row, col, None, COR_OBRIG)

        # ── Saidas ───────────────────────────────────────────────
        ws = wb.create_sheet("Saidas")
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A3"
        ws.row_dimensions[1].height = 34
        ws.merge_cells('A1:I1')
        c = ws.cell(row=1, column=1,
                    value='⚠ documento_origem = número do documento da ENTRADA que gerou este lote (ex: NF-0001). Deixe em branco para alocação FIFO automática.')
        c.font = Font(bold=True, name='Arial', size=10, color="7B3F00")
        c.fill = PatternFill("solid", fgColor=COR_INFO)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[2].height = 30
        cols_s = [("data",16,False),("documento",22,False),("cliente",28,False),
                  ("documento_origem",22,False),("produto",24,False),
                  ("quantidade",14,False),("unidade",12,False),("declaracao_fsc",20,False),("observacoes",30,False)]
        for i, (n, w, auto) in enumerate(cols_s, 1):
            hdr(ws, 2, i, n, w, auto=auto)
        ex_s = ["2026-03-20","NF-0010","Tramontina","NF-0001","Toras e Toretes",40,"t","FSC 100%","Baixa parcial"]
        bgs_s = [COR_OBRIG]*5 + [COR_OBRIG]*4
        for i, (v, bg) in enumerate(zip(ex_s, bgs_s), 1):
            cell(ws, 3, i, v, bg, italic=(bg == COR_AUTO))
        for row in range(4, 52):
            ws.row_dimensions[row].height = 18
            for col in range(1, 10):
                cell(ws, row, col, None, COR_OBRIG)

        # ── Transformacoes ───────────────────────────────────────
        wt = wb.create_sheet("Transformacoes")
        wt.sheet_view.showGridLines = False
        wt.freeze_panes = "A3"
        wt.row_dimensions[1].height = 34
        wt.merge_cells('A1:I1')
        c = wt.cell(row=1, column=1,
                    value='⚠ documento_origem = número do documento da ENTRADA de origem (ex: NF-0001). produto_origem é identificado automaticamente pelo sistema a partir deste campo.')
        c.font = Font(bold=True, name='Arial', size=10, color="7B3F00")
        c.fill = PatternFill("solid", fgColor=COR_INFO)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        wt.row_dimensions[2].height = 30
        cols_t = [("data",16,False),("documento",22,False),("cliente_final",28,False),
                  ("documento_origem",22,False),("produto_origem",24,True),
                  ("produto_destino",24,False),("quantidade_produzida",16,False),
                  ("unidade_destino",14,False),("observacoes",30,False)]
        for i, (n, w, auto) in enumerate(cols_t, 1):
            hdr(wt, 2, i, n, w, auto=auto)
        ex_t = ["2026-03-25","OP-0001","Tramontina","NF-0001","← sistema identifica","Madeira Serrada",12,"m3","Serrado a partir de toras"]
        bgs_t = [COR_OBRIG]*4 + [COR_AUTO] + [COR_OBRIG]*4
        for i, (v, bg) in enumerate(zip(ex_t, bgs_t), 1):
            cell(wt, 3, i, v, bg, italic=(bg == COR_AUTO))
        for row in range(4, 52):
            wt.row_dimensions[row].height = 18
            for col, bg in enumerate([COR_OBRIG]*4 + [COR_AUTO] + [COR_OBRIG]*4, 1):
                cell(wt, row, col, None, bg)

        wb.active = wi
        buffer = BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()
        cache.set(self.CACHE_KEY, xlsx_bytes, self.CACHE_TTL)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=modelo_importacao_fsc.xlsx'
        response.write(xlsx_bytes)
        return response


class ImportErrorsDownloadView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        job = _get_visible_import_job(request, kwargs.get('job_id'))
        if not job or not job.error_messages:
            messages.warning(request, 'Nenhum relatório de erros disponível para este job.')
            return redirect('report_import_workbook')

        wb = build_import_error_workbook(job.error_messages)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'erros_importacao_job_{job.pk}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


def _get_import_jobs_queryset(request):
    qs = ImportJob.objects.select_related('participant', 'created_by').order_by('-created_at', '-id')
    if request.user.is_manager or request.user.is_auditor:
        current_org = getattr(request.user, 'current_organization', None)
        if current_org:
            qs = qs.filter(participant__organization=current_org)
        return qs
    if getattr(request.user, 'participant_id', None):
        return qs.filter(participant=request.user.participant)
    return qs.none()


def _get_visible_import_job(request, job_id):
    if not job_id:
        return None
    return _get_import_jobs_queryset(request).filter(pk=job_id).first()



class ImportWorkbookView(LoginRequiredMixin, View):
    template_name = 'reports/import_workbook.html'

    def _build_form(self, request, *args, **kwargs):
        form = ImportWorkbookForm(*args, **kwargs)
        if request.user.is_manager:
            current_org = getattr(request.user, 'current_organization', None)
            if current_org:
                form.fields['participant'].queryset = form.fields['participant'].queryset.filter(organization=current_org)
        else:
            form.fields.pop('participant', None)
        return form

    def _render(self, request, form, preview=None, errors=None, summary=None, selected_job=None, validated_token=None, validated_filename=None):
        jobs = list(_get_import_jobs_queryset(request)[:15])
        if not selected_job and jobs:
            requested_job_id = request.GET.get('job')
            selected_job = _get_visible_import_job(request, requested_job_id) if requested_job_id else jobs[0]
        return render(request, self.template_name, {
            'form': form,
            'preview': preview or {},
            'preview_errors': humanize_import_errors(errors or []),
            'summary': summary or {},
            'jobs': jobs,
            'selected_job': selected_job,
            'selected_job_error_url': reverse('report_import_errors_download', kwargs={'job_id': selected_job.pk}) if selected_job and selected_job.error_messages else '',
            'is_job_running': bool(selected_job and selected_job.status in [ImportJob.STATUS_PENDING, ImportJob.STATUS_PROCESSING]),
            'import_mode': getattr(settings, 'IMPORT_MODE', 'sync'),
            'validated_token': validated_token,
            'validated_filename': validated_filename,
        })

    def get(self, request, *args, **kwargs):
        form = self._build_form(request, initial={'action': 'validate'})
        selected_job = _get_visible_import_job(request, request.GET.get('job'))
        return self._render(request, form, selected_job=selected_job)

    def _save_temp_file(self, request, uploaded_file):
        """Salva o arquivo em temp e guarda o caminho na sessão. Retorna o token."""
        token = str(uuid.uuid4())
        uploaded_file.seek(0)
        content = uploaded_file.read()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', prefix=f'import_{token}_')
        tmp.write(content)
        tmp.close()
        session_key = f'import_tmp_{token}'
        request.session[session_key] = {
            'path': tmp.name,
            'filename': getattr(uploaded_file, 'name', 'planilha.xlsx'),
            'participant_id': None,  # preenchido pelo chamador
        }
        return token

    def _load_temp_file(self, request, token):
        """Recupera o arquivo temporário da sessão. Retorna (bytes, filename) ou (None, None)."""
        session_key = f'import_tmp_{token}'
        data = request.session.get(session_key)
        if not data:
            return None, None
        path = data.get('path')
        filename = data.get('filename', 'planilha.xlsx')
        if not path or not os.path.exists(path):
            return None, None
        with open(path, 'rb') as f:
            content = f.read()
        return content, filename

    def _cleanup_temp_file(self, request, token):
        """Remove arquivo temp e limpa a sessão."""
        session_key = f'import_tmp_{token}'
        data = request.session.pop(session_key, None)
        if data and data.get('path'):
            try:
                os.unlink(data['path'])
            except OSError:
                pass

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action') or 'validate'

        # ── Confirmação sem re-upload ─────────────────────────────
        if action == 'confirm':
            token = request.POST.get('validated_token', '')
            participant_id = request.POST.get('participant_id')
            file_bytes, filename = self._load_temp_file(request, token)
            if not file_bytes:
                messages.error(request, 'Sessão expirada. Por favor, valide a planilha novamente.')
                form = self._build_form(request)
                return self._render(request, form)

            from participants.models import Participant as _Participant
            participant = _Participant.objects.filter(pk=participant_id).first() if participant_id else getattr(request.user, 'participant', None)
            if not participant:
                messages.error(request, 'Participante não encontrado.')
                form = self._build_form(request)
                return self._render(request, form)

            from io import BytesIO as _BytesIO
            import django.core.files.uploadedfile as _uf
            tmp_file = _uf.InMemoryUploadedFile(
                _BytesIO(file_bytes), 'workbook', filename,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                len(file_bytes), None
            )
            self._cleanup_temp_file(request, token)
            return self._do_import(request, participant, tmp_file, filename)

        # ── Upload normal ─────────────────────────────────────────
        form = self._build_form(request, request.POST, request.FILES)
        if not form.is_valid():
            return self._render(request, form)

        participant = form.cleaned_data.get('participant') if request.user.is_manager else request.user.participant
        if not participant:
            messages.error(request, 'Selecione um participante.')
            return self._render(request, form)

        uploaded_file = form.cleaned_data['workbook']
        workbook = openpyxl.load_workbook(uploaded_file)
        summary, errors, preview = build_import_preview(workbook, participant, request.user, persist=False)

        if action == 'validate' or errors:
            validated_token = None
            validated_filename = None
            if not errors:
                # Salva arquivo em temp para evitar re-upload na confirmação
                validated_token = self._save_temp_file(request, uploaded_file)
                validated_filename = getattr(uploaded_file, 'name', 'planilha.xlsx')
                request.session[f'import_tmp_{validated_token}']['participant_id'] = participant.pk
                request.session.modified = True
            if errors:
                messages.warning(request, f'Validação concluída com {len(errors)} inconsistências. Corrija a planilha antes de importar.')
            else:
                messages.success(request, 'Validação concluída sem inconsistências. Confirme para importar.')
            return self._render(request, form, preview=preview, errors=errors, summary=summary,
                                validated_token=validated_token, validated_filename=validated_filename)

        return self._do_import(request, participant, uploaded_file, getattr(uploaded_file, 'name', 'planilha.xlsx'))

    def _do_import(self, request, participant, uploaded_file, filename):
        """Executa a importação real (sync ou async)."""
        import_mode = getattr(settings, 'IMPORT_MODE', 'sync')

        if import_mode == 'sync':
            uploaded_file.seek(0)
            workbook = openpyxl.load_workbook(uploaded_file)
            summary, errors, preview = build_import_preview(workbook, participant, request.user, persist=True)
            if errors:
                messages.error(request, 'A importação encontrou inconsistências e não foi concluída. Revise a planilha.')
                form = self._build_form(request)
                return self._render(request, form, preview=preview, errors=errors, summary=summary)
            messages.success(
                request,
                f"Importação concluída. Entradas: {summary.get('entries', 0)}, saídas: {summary.get('sales', 0)}, transformações: {summary.get('transformations', 0)}."
            )
            return redirect('dashboard')

        uploaded_file.seek(0)
        workbook = openpyxl.load_workbook(uploaded_file)
        summary, errors, preview = build_import_preview(workbook, participant, request.user, persist=False)
        total_rows = count_workbook_rows(workbook)
        uploaded_file.seek(0)
        job = ImportJob.objects.create(
            participant=participant,
            created_by=request.user,
            workbook=uploaded_file,
            original_filename=filename,
            status=ImportJob.STATUS_PENDING,
            summary=_serialize_json_safe(summary),
            preview=_serialize_json_safe(preview),
            error_messages=[],
            progress_current=0,
            progress_total=total_rows,
        )
        messages.success(request, f'Importação enviada para a fila com sucesso. Job #{job.pk} criado para {filename or "planilha"}.')
        return redirect(f"{reverse('report_import_workbook')}?job={job.pk}")

