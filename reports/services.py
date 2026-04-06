
from decimal import Decimal
from io import BytesIO
from unicodedata import normalize

import openpyxl
from django.db import transaction
from django.utils import timezone

from catalog.models import Counterparty, Product
from transactions.models import EntryRecord, SaleRecord, TraceLot, TransformationRecord
from transactions.services import (
    convert_to_base,
    get_transformation_rule,
    reallocate_sale,
    reallocate_transformation_sources,
    sync_entry_lot,
    sync_transformation_target_lot,
)

from .models import ImportJob


def safe_str(value):
    return '' if value is None else str(value).strip()


def normalize_date(value):
    if hasattr(value, 'date'):
        return value.date()
    return value


def decimal_value(value):
    return Decimal(str(value)) if value not in (None, '') else Decimal('0')


def normalize_header(value):
    text = normalize('NFKD', safe_str(value)).encode('ascii', 'ignore').decode('ascii')
    return text.lower().replace(' ', '_')


def sheet_rows(sheet):
    # Detecta a linha de header buscando a primeira que contenha 'data'.
    # Isso permite que abas com linha de instrução na linha 1 (Saidas,
    # Transformacoes) funcionem corretamente — o header real fica na linha 2.
    headers = []
    header_row = 1
    for candidate in (1, 2):
        row_values = list(next(sheet.iter_rows(min_row=candidate, max_row=candidate, values_only=True), []))
        normalized = [normalize_header(cell) for cell in row_values]
        if 'data' in normalized:
            headers = normalized
            header_row = candidate
            break

    data_start = header_row + 1
    for idx, row in enumerate(sheet.iter_rows(min_row=data_start, values_only=True), start=data_start):
        values = list(row)
        if not any(value not in (None, '') for value in values):
            continue
        payload = {}
        for pos, header in enumerate(headers):
            if header:
                payload[header] = values[pos] if pos < len(values) else None
        yield idx, payload


def first_present(row, *keys):
    for key in keys:
        if key in row:
            return row.get(key)
    return None


def serialize_for_json(value):
    if value is None:
        return None
    if hasattr(value, 'isoformat'):
        try:
            return value.isoformat()
        except Exception:
            pass
    if isinstance(value, Decimal):
        return str(value)
    return value


def serialize_row_payload(row):
    return {key: serialize_for_json(value) for key, value in (row or {}).items()}


def serialize_payload_for_json(value):
    if isinstance(value, dict):
        return {key: serialize_payload_for_json(val) for key, val in value.items()}
    if isinstance(value, (list, tuple)):
        return [serialize_payload_for_json(item) for item in value]
    return serialize_for_json(value)


def make_import_error(sheet, row_number, message, row_data=None, field=''):
    return {
        'sheet': sheet,
        'row_number': row_number,
        'field': field or '',
        'message': str(message),
        'row_data': serialize_row_payload(row_data or {}),
    }


def humanize_import_errors(errors):
    messages_list = []
    for error in errors:
        if isinstance(error, str):
            messages_list.append(error)
            continue
        sheet = error.get('sheet', 'Planilha')
        row_number = error.get('row_number', '?')
        field = error.get('field')
        message = error.get('message', 'Erro não identificado.')
        prefix = f'{sheet} linha {row_number}'
        if field:
            prefix += f' [{field}]'
        messages_list.append(f'{prefix}: {message}')
    return messages_list


def build_import_error_workbook(errors):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Erros'
    ws.append(['Aba', 'Linha', 'Campo', 'Erro', 'Dados da linha'])
    for error in errors:
        if isinstance(error, str):
            ws.append(['', '', '', error, ''])
            continue
        row_data = error.get('row_data') or {}
        raw_text = '; '.join(f'{key}={value}' for key, value in row_data.items())
        ws.append([
            error.get('sheet', ''),
            error.get('row_number', ''),
            error.get('field', ''),
            error.get('message', ''),
            raw_text,
        ])
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 60)
    return wb


def count_workbook_rows(workbook):
    total = 0
    for sheet_name in ('Entradas', 'Saidas', 'Transformacoes'):
        if sheet_name in workbook.sheetnames:
            total += sum(1 for _idx, _row in sheet_rows(workbook[sheet_name]))
    return total


def _coerce_product(product_name, *, default_unit=None, create_if_missing=False):
    product_name = safe_str(product_name)
    if not product_name:
        raise ValueError('Produto não informado.')
    if create_if_missing:
        product, _ = Product.objects.get_or_create(
            name=product_name,
            defaults={'unit': default_unit or 'm3', 'active': True},
        )
        return product
    product = Product.objects.filter(name=product_name).first()
    if product is None:
        raise ValueError(f'Produto "{product_name}" não encontrado. Cadastre-o antes de importar.')
    return product


# CORREÇÃO: adicionado parâmetro create_if_missing (padrão False).
# Na fase de preview (persist=False), apenas buscamos a contraparte sem criá-la,
# evitando efeitos colaterais no banco durante a validação.
def _get_or_create_counterparty(participant, name, *, type_, create_if_missing=True):
    normalized_name = safe_str(name)
    if not normalized_name:
        return None
    if create_if_missing:
        return Counterparty.objects.get_or_create(
            participant=participant,
            name=normalized_name,
            defaults={'type': type_},
        )[0]
    return Counterparty.objects.filter(participant=participant, name=normalized_name).first()


def _resolve_preferred_lot(participant, documento_origem, product=None):
    """
    Busca o TraceLot pelo número do documento da entrada de origem.
    Aceita tanto o número do documento (ex: 'NF-0001') quanto o código do lote.
    Retorna o TraceLot encontrado ou None se não informado.
    Levanta ValueError se o documento foi informado mas não foi encontrado.
    """
    ref = safe_str(documento_origem)
    if not ref:
        return None

    qs = TraceLot.objects.filter(participant=participant, source_type='entry')
    if product:
        qs = qs.filter(product=product)

    # Tenta por número do documento da entrada
    lot = qs.filter(entry__document_number=ref).first()
    if lot:
        return lot

    # Tenta por código do lote da entrada
    lot = qs.filter(entry__batch_code=ref).first()
    if lot:
        return lot

    raise ValueError(
        f'documento_origem "{ref}" não encontrado. '
        f'Verifique se o número do documento da entrada foi digitado corretamente '
        f'e se a entrada já foi importada.'
    )


def build_import_preview(workbook, participant, user, persist=False):
    """
    Processa a planilha em duas fases quando persist=True:
      Fase 1 — Entradas: grava todas as entradas e seus lotes.
      Fase 2 — Saídas e Transformações: agora encontram os lotes da fase 1.
    No modo preview (persist=False) tudo roda sem gravar, apenas validando.
    """
    summary = {'entries': 0, 'sales': 0, 'transformations': 0}
    errors = []
    previews = {'entries': [], 'sales': [], 'transformations': []}

    if 'Entradas' in workbook.sheetnames:
        sheet = workbook['Entradas']
        for idx, row in sheet_rows(sheet):
            try:
                data = first_present(row, 'data')
                if not data:
                    raise ValueError('Data não informada.')
                documento = first_present(row, 'documento')
                fornecedor_nome = first_present(row, 'fornecedor')
                produto_nome = first_present(row, 'produto')
                quantidade = first_present(row, 'quantidade')
                unidade = first_present(row, 'unidade')
                declaracao = first_present(row, 'declaracao_fsc')
                lote = first_present(row, 'lote')
                observacoes = first_present(row, 'observacoes')

                unidade = safe_str(unidade) or 'm3'
                # CORREÇÃO: só cria produto novo durante a gravação real (persist=True).
                product = _coerce_product(produto_nome, default_unit=unidade, create_if_missing=persist)
                supplier_name = safe_str(fornecedor_nome) or 'Não informado'
                # CORREÇÃO: só cria fornecedor durante a gravação real (persist=True).
                supplier = _get_or_create_counterparty(participant, supplier_name, type_='supplier', create_if_missing=persist)
                quantidade = decimal_value(quantidade)
                quantity_base = convert_to_base(product, quantidade, unidade)
                preview = {
                    'linha': idx,
                    'data': data,
                    'documento': safe_str(documento),
                    'supplier': supplier_name,
                    'product': product.name,
                    'quantity': quantidade,
                    'unit': unidade,
                    'quantity_base': quantity_base,
                }
                previews['entries'].append(preview)
                summary['entries'] += 1
                if persist:
                    obj = EntryRecord.objects.create(
                        participant=participant,
                        movement_date=normalize_date(data),
                        document_number=safe_str(documento),
                        supplier=supplier,
                        product=product,
                        quantity=quantidade,
                        movement_unit=unidade,
                        unit_snapshot=product.unit,
                        quantity_base=quantity_base,
                        fsc_claim=safe_str(declaracao),
                        batch_code=safe_str(lote),
                        notes=safe_str(observacoes),
                        created_by=user,
                        status='submitted',
                    )
                    sync_entry_lot(obj)
            except Exception as exc:
                errors.append(make_import_error('Entradas', idx, exc, row))

    if 'Saidas' in workbook.sheetnames:
        sheet = workbook['Saidas']
        for idx, row in sheet_rows(sheet):
            try:
                data = first_present(row, 'data')
                if not data:
                    raise ValueError('Data não informada.')
                documento = first_present(row, 'documento')
                cliente_nome = first_present(row, 'cliente')
                produto_nome = first_present(row, 'produto')
                quantidade = first_present(row, 'quantidade')
                unidade = first_present(row, 'unidade')
                declaracao = first_present(row, 'declaracao_fsc')
                lote = first_present(row, 'lote')
                observacoes = first_present(row, 'observacoes')
                documento_origem = safe_str(first_present(row, 'documento_origem'))

                # Se produto vier com texto informativo do modelo (ex: "← sistema identifica"),
                # ignora e deriva o produto a partir do documento_origem.
                # Produto é sempre informado diretamente na saída (igual à entrada).
                # A coluna produto no modelo é obrigatória e azul.
                product = _coerce_product(safe_str(produto_nome))

                customer_name = safe_str(cliente_nome)
                if not customer_name:
                    raise ValueError('Cliente não informado.')
                customer = _get_or_create_counterparty(participant, customer_name, type_='customer', create_if_missing=persist)
                quantidade = decimal_value(quantidade)

                unidade = safe_str(unidade) or product.unit
                quantity_base = convert_to_base(product, quantidade, unidade)

                # Resolve lote preferencial via documento_origem ou id_lote_origem (legado).
                preferred_lot = _resolve_preferred_lot(participant, documento_origem, product=product) if persist else None
                if not preferred_lot and persist:
                    lot_id = safe_str(first_present(row, 'id_lote_origem'))
                    if lot_id:
                        preferred_lot = TraceLot.objects.get(pk=lot_id, participant=participant, product=product)

                preview = {
                    'linha': idx,
                    'data': data,
                    'documento': safe_str(documento),
                    'customer': customer_name,
                    'product': product.name,
                    'quantity': quantidade,
                    'unit': unidade,
                    'quantity_base': quantity_base,
                    'documento_origem': documento_origem or '(FIFO automático)',
                }
                previews['sales'].append(preview)
                summary['sales'] += 1
                if persist:
                    obj = SaleRecord.objects.create(
                        participant=participant,
                        movement_date=normalize_date(data),
                        document_number=safe_str(documento),
                        customer=customer,
                        product=product,
                        quantity=quantidade,
                        movement_unit=unidade,
                        unit_snapshot=product.unit,
                        quantity_base=quantity_base,
                        fsc_claim=safe_str(declaracao),
                        batch_code=safe_str(lote),
                        notes=safe_str(observacoes),
                        created_by=user,
                        status='submitted',
                    )
                    reallocate_sale(obj, preferred_lot=preferred_lot)
            except Exception as exc:
                errors.append(make_import_error('Saidas', idx, exc, row))

    if 'Transformacoes' in workbook.sheetnames:
        sheet = workbook['Transformacoes']
        for idx, row in sheet_rows(sheet):
            try:
                data = first_present(row, 'data')
                if not data:
                    raise ValueError('Data não informada.')
                documento = first_present(row, 'documento')
                cliente_nome = first_present(row, 'cliente_final', 'cliente')
                source_product_name = first_present(row, 'produto_origem')
                target_product_name = first_present(row, 'produto_destino')
                target_quantity = first_present(row, 'quantidade_produzida', 'quantidade_destino', 'quantidade')
                target_unit = first_present(row, 'unidade_destino', 'unidade')
                observacoes = first_present(row, 'observacoes')

                documento_origem = safe_str(first_present(row, 'documento_origem'))

                # Se documento_origem informado, deriva produto_origem do lote automaticamente.
                # Caso contrário, exige que produto_origem seja preenchido na planilha.
                if documento_origem and persist:
                    lot_for_source = _resolve_preferred_lot(participant, documento_origem)
                    if lot_for_source:
                        source_product = lot_for_source.product
                    else:
                        source_product = _coerce_product(source_product_name)
                elif safe_str(source_product_name) and safe_str(source_product_name) != '← preenchido pelo sistema':
                    source_product = _coerce_product(source_product_name)
                else:
                    raise ValueError('Informe produto_origem ou documento_origem para identificar a origem da transformação.')

                # CORREÇÃO: só cria produto destino durante a gravação real (persist=True).
                target_product = _coerce_product(target_product_name, default_unit=safe_str(target_unit) or 'm3', create_if_missing=persist)
                rule = get_transformation_rule(source_product, target_product, participant=participant)
                if not rule:
                    raise ValueError('Não existe regra de transformação cadastrada para os produtos selecionados para esta empresa.')

                target_quantity = decimal_value(target_quantity)
                target_unit = safe_str(target_unit) or target_product.unit
                target_quantity_base = convert_to_base(target_product, target_quantity, target_unit)
                if not rule.yield_factor:
                    raise ValueError('A regra de transformação está com fator de rendimento inválido.')
                source_quantity_base = (target_quantity_base / Decimal(str(rule.yield_factor))).quantize(Decimal('0.001'))

                preview = {
                    'linha': idx,
                    'data': data,
                    'document_number': safe_str(documento),
                    'customer': safe_str(cliente_nome),
                    'documento_origem': documento_origem or '(FIFO automático)',
                    'source_product': source_product.name,
                    'source_quantity_base': source_quantity_base,
                    'source_unit': source_product.unit,
                    'target_product': target_product.name,
                    'target_quantity': target_quantity,
                    'target_unit': target_unit,
                    'target_quantity_base': target_quantity_base,
                    'yield_factor': rule.yield_factor,
                }
                previews['transformations'].append(preview)
                summary['transformations'] += 1
                if persist:
                    preferred_lot = _resolve_preferred_lot(participant, documento_origem, product=source_product)
                    if not preferred_lot:
                        lot_id = safe_str(first_present(row, 'id_lote_origem'))
                        if lot_id:
                            preferred_lot = TraceLot.objects.get(pk=lot_id, participant=participant, product=source_product)
                    # CORREÇÃO: só cria cliente durante a gravação real (persist=True).
                    customer = _get_or_create_counterparty(participant, cliente_nome, type_='customer', create_if_missing=True) if safe_str(cliente_nome) else None
                    obj = TransformationRecord.objects.create(
                        participant=participant,
                        movement_date=normalize_date(data),
                        document_number=safe_str(documento),
                        customer=customer,
                        source_product=source_product,
                        source_quantity=source_quantity_base,
                        source_unit=source_product.unit,
                        source_quantity_base=source_quantity_base,
                        target_product=target_product,
                        target_quantity_base=target_quantity_base,
                        target_unit_snapshot=target_product.unit,
                        yield_factor_snapshot=rule.yield_factor,
                        notes=safe_str(observacoes),
                        created_by=user,
                    )
                    reallocate_transformation_sources(obj, preferred_lot=preferred_lot)
                    sync_transformation_target_lot(obj)
            except Exception as exc:
                errors.append(make_import_error('Transformacoes', idx, exc, row))

    return summary, errors, previews


def _set_job_progress(job, *, current=None, total=None, status=None):
    fields = ['updated_at']
    if current is not None:
        job.progress_current = current
        fields.append('progress_current')
    if total is not None:
        job.progress_total = total
        fields.append('progress_total')
    if status is not None:
        job.status = status
        fields.append('status')
    job.last_heartbeat = timezone.now()
    fields.append('last_heartbeat')
    job.save(update_fields=fields)



# Limite máximo de linhas por planilha no modo sync (sem worker).
# Protege contra timeout no Render free tier (~120s).
# Em modo async com worker, o limite pode ser maior sem risco.
IMPORT_MAX_ROWS_SYNC = 500
IMPORT_MAX_ROWS_ASYNC = 5000


def check_row_limit(total_rows, *, async_mode=False):
    """
    Levanta ValueError se a planilha exceder o limite permitido para o modo atual.
    Deve ser chamado antes de qualquer processamento.
    """
    limit = IMPORT_MAX_ROWS_ASYNC if async_mode else IMPORT_MAX_ROWS_SYNC
    if total_rows > limit:
        modo = 'assíncrono' if async_mode else 'síncrono'
        raise ValueError(
            f'A planilha contém {total_rows} linhas, mas o limite no modo {modo} é {limit}. '
            f'Divida o arquivo em partes menores e importe separadamente.'
        )


def _process_entries_only(workbook, participant, user):
    """Fase 1: processa apenas a aba Entradas com persist=True."""
    summary = {'entries': 0, 'sales': 0, 'transformations': 0}
    errors = []
    previews = {'entries': [], 'sales': [], 'transformations': []}

    if 'Entradas' not in workbook.sheetnames:
        return summary, errors, previews

    sheet = workbook['Entradas']
    for idx, row in sheet_rows(sheet):
        try:
            data = first_present(row, 'data')
            if not data:
                raise ValueError('Data não informada.')
            documento = first_present(row, 'documento')
            fornecedor_nome = first_present(row, 'fornecedor')
            produto_nome = first_present(row, 'produto')
            quantidade = first_present(row, 'quantidade')
            unidade = first_present(row, 'unidade')
            declaracao = first_present(row, 'declaracao_fsc')
            lote = first_present(row, 'lote')
            observacoes = first_present(row, 'observacoes')

            unidade = safe_str(unidade) or 'm3'
            product = _coerce_product(produto_nome, default_unit=unidade, create_if_missing=True)
            supplier_name = safe_str(fornecedor_nome) or 'Não informado'
            supplier = _get_or_create_counterparty(participant, supplier_name, type_='supplier', create_if_missing=True)
            quantidade = decimal_value(quantidade)
            quantity_base = convert_to_base(product, quantidade, unidade)
            previews['entries'].append({
                'linha': idx, 'data': data, 'documento': safe_str(documento),
                'supplier': supplier_name, 'product': product.name,
                'quantity': quantidade, 'unit': unidade, 'quantity_base': quantity_base,
            })
            summary['entries'] += 1
            obj = EntryRecord.objects.create(
                participant=participant,
                movement_date=normalize_date(data),
                document_number=safe_str(documento),
                supplier=supplier,
                product=product,
                quantity=quantidade,
                movement_unit=unidade,
                unit_snapshot=product.unit,
                quantity_base=quantity_base,
                fsc_claim=safe_str(declaracao),
                batch_code=safe_str(lote),
                notes=safe_str(observacoes),
                created_by=user,
                status='submitted',
            )
            sync_entry_lot(obj)
        except Exception as exc:
            errors.append(make_import_error('Entradas', idx, exc, row))

    return summary, errors, previews


def _process_sales_and_transformations(workbook, participant, user):
    """Fase 2: processa Saidas e Transformacoes com persist=True.
    As entradas já foram gravadas na Fase 1, então os lotes já existem.
    """
    summary = {'entries': 0, 'sales': 0, 'transformations': 0}
    errors = []
    previews = {'entries': [], 'sales': [], 'transformations': []}

    # Reutiliza build_import_preview só para saídas e transformações
    # passando um workbook sem a aba Entradas para evitar reprocessamento.
    from openpyxl import Workbook as _Workbook
    wb_partial = _Workbook()
    wb_partial.remove(wb_partial.active)

    for sheet_name in ('Saidas', 'Transformacoes'):
        if sheet_name in workbook.sheetnames:
            # Copia a aba para o workbook parcial
            from openpyxl.utils import get_column_letter
            src = workbook[sheet_name]
            dst = wb_partial.create_sheet(sheet_name)
            for row in src.iter_rows(values_only=True):
                dst.append(list(row))

    partial_summary, partial_errors, partial_previews = build_import_preview(
        wb_partial, participant, user, persist=True
    )
    summary['sales'] = partial_summary['sales']
    summary['transformations'] = partial_summary['transformations']
    previews['sales'] = partial_previews.get('sales', [])
    previews['transformations'] = partial_previews.get('transformations', [])
    errors = partial_errors

    return summary, errors, previews


def process_import_job(job):
    job.workbook.open('rb')
    try:
        payload = job.workbook.read()
    finally:
        job.workbook.close()

    workbook = openpyxl.load_workbook(BytesIO(payload))
    total_rows = count_workbook_rows(workbook)

    # Verifica limite de linhas antes de qualquer processamento.
    # async_mode=True quando há worker rodando (IMPORT_MODE=async).
    from django.conf import settings
    async_mode = getattr(settings, 'IMPORT_MODE', 'sync') == 'async'
    try:
        check_row_limit(total_rows, async_mode=async_mode)
    except ValueError as exc:
        job.status = ImportJob.STATUS_FAILED
        job.error_messages = [make_import_error('Planilha', '', exc)]
        job.finished_at = timezone.now()
        job.last_heartbeat = timezone.now()
        job.save(update_fields=['status', 'error_messages', 'finished_at', 'last_heartbeat', 'updated_at'])
        return job

    preview_summary, preview_errors, preview = build_import_preview(workbook, job.participant, job.created_by, persist=False)

    job.summary = serialize_payload_for_json(preview_summary)
    job.preview = serialize_payload_for_json(preview)
    job.progress_total = total_rows
    job.progress_current = 0
    job.last_heartbeat = timezone.now()
    job.save(update_fields=['summary', 'preview', 'progress_total', 'progress_current', 'last_heartbeat', 'updated_at'])

    if preview_errors:
        job.status = ImportJob.STATUS_FAILED
        job.error_messages = serialize_payload_for_json(preview_errors)
        job.finished_at = timezone.now()
        job.last_heartbeat = timezone.now()
        job.save(update_fields=['status', 'error_messages', 'finished_at', 'last_heartbeat', 'updated_at'])
        return job

    workbook = openpyxl.load_workbook(BytesIO(payload))
    try:
        with transaction.atomic():
            # Fase 1 — processa APENAS as entradas e grava os lotes.
            # Isso garante que saídas e transformações encontrem os lotes
            # mesmo quando entradas e saídas estão na mesma planilha.
            phase1_summary, phase1_errors, phase1_preview = _process_entries_only(
                workbook, job.participant, job.created_by
            )
            if phase1_errors:
                raise ValueError(humanize_import_errors(phase1_errors)[0])

            # Fase 2 — processa saídas e transformações (lotes já existem no banco).
            phase2_summary, phase2_errors, phase2_preview = _process_sales_and_transformations(
                workbook, job.participant, job.created_by
            )
            if phase2_errors:
                raise ValueError(humanize_import_errors(phase2_errors)[0])

            persisted_summary = {
                'entries': phase1_summary['entries'],
                'sales': phase2_summary['sales'],
                'transformations': phase2_summary['transformations'],
            }
            persisted_preview = {
                'entries': phase1_preview['entries'],
                'sales': phase2_preview['sales'],
                'transformations': phase2_preview['transformations'],
            }
            persisted_errors = []
    except Exception as exc:
        job.status = ImportJob.STATUS_FAILED
        job.error_messages = serialize_payload_for_json(preview_errors or [make_import_error('Importação', '', exc)])
        job.finished_at = timezone.now()
        job.last_heartbeat = timezone.now()
        job.save(update_fields=['status', 'error_messages', 'finished_at', 'last_heartbeat', 'updated_at'])
        raise

    job.status = ImportJob.STATUS_COMPLETED
    job.summary = serialize_payload_for_json(persisted_summary)
    job.preview = serialize_payload_for_json(persisted_preview)
    job.error_messages = []
    job.progress_current = total_rows
    job.finished_at = timezone.now()
    job.last_heartbeat = timezone.now()
    job.save(update_fields=['status', 'summary', 'preview', 'error_messages', 'progress_current', 'finished_at', 'last_heartbeat', 'updated_at'])
    return job
